#!/usr/bin/env python3
"""
Calcula a pontuação do Conclave MR a partir de:
  • planilha legada (aba «Entrada», 12 colunas de medalhas por igreja), ou
  • arquivo de projeto JSON (evento + dados com pódio por prova).

Uso:
  python3 calcular_conclave.py
  python3 calcular_conclave.py -i Entrada_Conclave_MR.xlsx -o Resultado_Conclave_MR.xlsx \\
      --evento eventos/conclave-2026-1.evento.json
  python3 calcular_conclave.py --projeto projeto.json -o Resultado_Conclave_MR.xlsx
  python3 calcular_conclave.py --criar-template

Requisito: pip install -r requirements.txt
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from engine import classificacao_ordenada, compute_totals, empty_dados_template, validate_evento_minimal

# Pesos conforme tabela do regulamento (template legado; projeto JSON usa evento.pesos)
PESO = {
    "inscricao": 100,
    "pontualidade": 200,
    "uniforme": 50,
    "biblia": 50,
    "visitante": 10,
    "animacao": 150,
    "mau_comportamento": -150,
}

GINCANA = {
    "ou": 300,
    "pt": 200,
    "br": 100,
}

HEAD_ENTRADA = [
    "Igreja",
    "Inscrição (1=sim)",
    "Pontualidade (1=sim)",
    "Total MR",
    "MR c/ camisa",
    "MR c/ bíblia",
    "Visitantes",
    'Animação (x=sim)',
    'Mau comport. (x=sim)',
    "Esgrima Jun",
    "Esgrima Adl",
    "Esgrima Juv",
    "Debate Jun",
    "Debate Adl",
    "Debate Juv",
    "Esgr. Av. Jun",
    "Esgr. Av. Adl",
    "Esgr. Av. Juv",
    "Prova escr. Jun",
    "Prova escr. Adl",
    "Prova escr. Juv",
    "Pontuação Extra (manual)",
]


def _truthy_marcacao(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("x", "1", "sim", "s", "true", "verdadeiro", "ok", "✓")


def _num(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _pontos_participacao(row: dict) -> float:
    b = _num(row.get("inscricao"), 0)
    c = _num(row.get("pontualidade"), 0)
    d = int(_num(row.get("mr_total"), 0))
    e = int(_num(row.get("mr_camisa"), 0))
    f = int(_num(row.get("mr_biblia"), 0))
    g = int(_num(row.get("visitantes"), 0))
    anim = _truthy_marcacao(row.get("animacao"))

    if (b + d) <= 0:
        return 0.0

    total = b * PESO["inscricao"] + c * PESO["pontualidade"]
    if d > 0 and e == d:
        total += PESO["uniforme"]
    if d > 0 and f == d:
        total += PESO["biblia"]
    total += g * PESO["visitante"]
    if anim:
        total += PESO["animacao"]
    return total


def _pontos_punicoes(row: dict) -> float:
    p = 0.0
    if _truthy_marcacao(row.get("mau_comportamento")):
        p += PESO["mau_comportamento"]
    return p


def _pontos_gincana(cells: list) -> float:
    t = 0.0
    for v in cells:
        if v is None:
            continue
        s = str(v).strip().lower()
        if s in GINCANA:
            t += GINCANA[s]
    return t


def _ler_entrada(ws) -> list[dict]:
    """Espera linha 1 = cabeçalhos HEAD_ENTRADA."""
    rows_out = []
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if nome is None or str(nome).strip() == "":
            continue
        row = {
            "igreja": str(nome).strip(),
            "inscricao": ws.cell(r, 2).value,
            "pontualidade": ws.cell(r, 3).value,
            "mr_total": ws.cell(r, 4).value,
            "mr_camisa": ws.cell(r, 5).value,
            "mr_biblia": ws.cell(r, 6).value,
            "visitantes": ws.cell(r, 7).value,
            "animacao": ws.cell(r, 8).value,
            "mau_comportamento": ws.cell(r, 9).value,
            "gincana": [ws.cell(r, c).value for c in range(10, 22)],
            "embaixadas": ws.cell(r, 22).value,
        }
        rows_out.append(row)
    return rows_out


def _rank(totais: list[float]) -> list[int]:
    """Ranking 1,1,3… por total (empates na mesma posição)."""
    n = len(totais)
    out = [0] * n
    order = sorted(range(n), key=lambda i: totais[i], reverse=True)
    for j, i in enumerate(order):
        if j == 0:
            out[i] = 1
        else:
            prev = order[j - 1]
            if totais[i] == totais[prev]:
                out[i] = out[prev]
            else:
                out[i] = j + 1
    return out


def calcular(rows: list[dict]) -> tuple[list[dict], list[float]]:
    """Cálculo legado (sem JSON de evento): mesma ordem de linhas da planilha."""
    detalhes = []
    totais = []
    for row in rows:
        p_part = _pontos_participacao(row)
        p_puni = _pontos_punicoes(row)
        p_ginc = _pontos_gincana(row["gincana"])
        p_emb = _num(
            row.get("pontuacao_extra") if row.get("pontuacao_extra") is not None else row.get("embaixadas"),
            0,
        )
        total = p_part + p_puni + p_ginc + p_emb
        detalhes.append(
            {
                "igreja": row["igreja"],
                "participacao": p_part,
                "punicoes": p_puni,
                "gincana": p_ginc,
                "pontuacao_extra": p_emb,
                "total": total,
            }
        )
        totais.append(total)
    return detalhes, totais


def match_igreja_id(nome: str, igrejas: list[dict[str, Any]]) -> str | None:
    nome_clean = str(nome).strip().lower()
    for g in igrejas:
        if g["nome"].strip().lower() == nome_clean:
            return str(g["id"])
    for g in igrejas:
        gn = g["nome"].strip().lower()
        if nome_clean in gn or gn in nome_clean:
            return str(g["id"])
    return None


def normalize_medal_cell(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("ou", "ouro", "gold", "g"):
        return "ou"
    if s in ("pt", "prata", "silver", "p"):
        return "pt"
    if s in ("br", "bronze", "b"):
        return "br"
    return None


def legacy_rows_to_dados(rows: list[dict[str, Any]], evento: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    """
    Converte linhas da planilha legada (gincana = 12 células ou/pt/br) para dados.podium + participacao.
    Ordem das colunas de gincana = provas ordenadas por «ordem» no JSON do evento.
    """
    warnings: list[str] = []
    igrejas = evento["igrejas"]
    provas_ord = sorted(evento["provas"], key=lambda p: p.get("ordem", 0))
    prova_ids = [p["id"] for p in provas_ord]
    ncols = len(prova_ids)
    dados = empty_dados_template([g["id"] for g in igrejas], prova_ids)
    participacao = dados["participacao"]
    podium = dados["podium"]

    for row in rows:
        nome = row.get("igreja", "")
        gid = match_igreja_id(str(nome), igrejas)
        if not gid:
            warnings.append(f"Igreja não mapeada (ignorada): {nome!r}")
            continue
        participacao[gid] = {
            "inscricao": row.get("inscricao"),
            "pontualidade": row.get("pontualidade"),
            "mr_total": row.get("mr_total"),
            "mr_camisa": row.get("mr_camisa"),
            "mr_biblia": row.get("mr_biblia"),
            "visitantes": row.get("visitantes"),
            "animacao": row.get("animacao"),
            "mau_comportamento": row.get("mau_comportamento"),
            "pontuacao_extra": _num(
                row.get("pontuacao_extra") if row.get("pontuacao_extra") is not None else row.get("embaixadas"),
                0,
            ),
        }

    cells_list = [row.get("gincana") or [] for row in rows]
    for row in rows:
        gid = match_igreja_id(str(row.get("igreja", "")), igrejas)
        if not gid:
            continue
        cells = row.get("gincana") or []
        for idx in range(min(len(cells), ncols)):
            m = normalize_medal_cell(cells[idx])
            if not m:
                continue
            pid = prova_ids[idx]
            prev = podium[pid][m].get("igrejaId")
            if prev and prev != gid:
                warnings.append(f"Conflito na prova {pid} ({m}): já havia outra igreja; última prevalece.")
            podium[pid][m] = {"igrejaId": gid, "competidor": ""}

    return dados, warnings


def escrever_resultado(path_out: Path, out: dict[str, Any]) -> None:
    """Grava planilha na ordem da classificação (inclui critérios de desempate)."""
    ordem = classificacao_ordenada(
        out["detalhes"], out["ranks"], out.get("tiebreakByIgreja")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    headers = [
        "Igreja",
        "Participação",
        "Punições",
        "Gincana",
        "Pontuação Extra",
        "TOTAL",
        "Posição",
    ]
    ws.append(headers)
    for d in ordem:
        nome = d.get("igreja") or d.get("igrejaId", "")
        extra = d.get("pontuacao_extra")
        if extra is None:
            extra = d.get("embaixadas", 0)
        ws.append(
            [
                nome,
                round(d["participacao"], 2),
                round(d["punicoes"], 2),
                round(d["gincana"], 2),
                round(float(extra or 0), 2),
                round(d["total"], 2),
                d["posicao"],
            ]
        )

    for c in range(1, 8):
        ws.cell(1, c).font = Font(bold=True)
    ws.freeze_panes = "A2"
    wb.save(path_out)


def criar_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Entrada"

    fill = PatternFill("solid", fgColor="FFF9C4")
    for col, h in enumerate(HEAD_ENTRADA, start=1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")

    exemplo = [
        "Exemplo — apague ou substitua",
        1,
        1,
        4,
        4,
        4,
        0,
        "x",
        "",
        "ou",
        "",
        "",
        "pt",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        0,
    ]
    for col, v in enumerate(exemplo, start=1):
        cell = ws.cell(2, col, v)
        cell.fill = fill

    for col in range(1, len(HEAD_ENTRADA) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(22, 14 + (col == 1) * 10)

    inst = wb.create_sheet("Como usar", 1)
    inst["A1"] = "Como usar"
    inst["A1"].font = Font(bold=True, size=14)
    texto = [
        "",
        "1. Preencha APENAS a aba «Entrada» (linhas a partir da 2).",
        "2. Não use fórmulas nas células amarelas — só números ou códigos ou, pt, br (ouro, prata, bronze).",
        "3. Salve o arquivo e rode no Terminal:",
        "      python3 calcular_conclave.py -i Entrada_Conclave_MR.xlsx -o Resultado_Conclave_MR.xlsx --evento eventos/conclave-2026-1.evento.json",
        "4. Ou use o aplicativo web (index.html) com projeto JSON e exporte.",
        "",
        "Uniforme e bíblia: 50 pts cada somente se MR c/ camisa = Total MR e MR c/ bíblia = Total MR.",
        "Visitantes: 10 pontos por visitante.",
        "Punições: mau comport. = -150 (ajuste os pesos no JSON do evento se mudar o regulamento).",
    ]
    for i, line in enumerate(texto, start=2):
        inst.cell(i, 1, line)
    inst.column_dimensions["A"].width = 92

    wb.save(path)
    print("Template criado:", path)


def load_projeto(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if "evento" not in data or "dados" not in data:
        raise ValueError("Projeto JSON precisa ter chaves «evento» e «dados».")
    return data


def main() -> None:
    base = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(description="Calcula pontuação Conclave MR (planilha legada ou projeto JSON).")
    ap.add_argument("-i", "--entrada", type=Path, default=base / "Entrada_Conclave_MR.xlsx", help="Planilha de entrada (legado)")
    ap.add_argument("-o", "--saida", type=Path, default=base / "Resultado_Conclave_MR.xlsx", help="Arquivo de saída")
    ap.add_argument("--projeto", type=Path, help="Projeto JSON (evento + dados) — alternativa à planilha")
    ap.add_argument(
        "--evento",
        type=Path,
        default=base / "eventos" / "conclave-2026-1.evento.json",
        help="JSON de configuração do evento (obrigatório para converter planilha legada para o motor novo)",
    )
    ap.add_argument("--criar-template", action="store_true", help="Cria Entrada_Conclave_MR.xlsx modelo e sai")
    args = ap.parse_args()

    if args.criar_template:
        criar_template(args.entrada)
        return

    if args.projeto:
        if not args.projeto.exists():
            print("Projeto não encontrado:", args.projeto)
            raise SystemExit(1)
        projeto = load_projeto(args.projeto)
        evento = projeto["evento"]
        dados = projeto["dados"]
        errs = validate_evento_minimal(evento)
        if errs:
            print("Erros no evento:", errs)
            raise SystemExit(1)
        out = compute_totals(evento, dados)
        escrever_resultado(args.saida, out)
        print("Gerado:", args.saida, f"({len(detalhes)} igrejas) [projeto JSON]")
        return

    if not args.entrada.exists():
        print("Arquivo de entrada não encontrado:", args.entrada)
        print("Crie o modelo com: python3 calcular_conclave.py --criar-template")
        raise SystemExit(1)

    if not args.evento.exists():
        print("Arquivo de evento não encontrado:", args.evento)
        print("Indique --evento apontando para um .evento.json (ex.: eventos/conclave-2026-1.evento.json).")
        raise SystemExit(1)

    evento = json.loads(args.evento.read_text(encoding="utf-8"))
    errs = validate_evento_minimal(evento)
    if errs:
        print("Erros no evento:", errs)
        raise SystemExit(1)

    wb = load_workbook(args.entrada, data_only=True)
    if "Entrada" not in wb.sheetnames:
        print("A planilha precisa ter uma aba chamada «Entrada».")
        raise SystemExit(1)
    rows = _ler_entrada(wb["Entrada"])
    if not rows:
        print("Nenhuma igreja na coluna A (linhas 2 em diante).")
        raise SystemExit(1)

    dados, warnings = legacy_rows_to_dados(rows, evento)
    for w in warnings:
        print("Aviso:", w)

    out = compute_totals(evento, dados)
    escrever_resultado(args.saida, out)
    print("Gerado:", args.saida, f"({len(detalhes)} igrejas) [motor + planilha legada]")


if __name__ == "__main__":
    main()
